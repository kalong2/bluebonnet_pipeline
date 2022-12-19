from openpyxl import load_workbook
import sys
import re
import os

base_dir = "/work/NBS/"
print("start fail script")
#Grab template excel template, set worksheet
#workbook_path = base_dir + "/excel_parse/variant_worksheet_v2.xlsx"
#wb = load_workbook(workbook_path)
#ws = wb.active

gene=sys.argv[2]
condition = {
  "ABCD1": "X-ALD",
  "HBB": "Hemoglobinopathy",
  "ACADVL": "VLCADD"
}
website = {
  "ABCD1": "https://adrenoleukodystrophy.info/mutations-and-variants-in-abcd1",
  "HBB": "https://globin.bx.psu.edu/cgi-bin/hbvar/query_vars3",
  "ACADVL": "N/A"
}


#$prefix $rsid $gene $var $locus $pos $end $chrom
var=sys.argv[3]
locus=sys.argv[4]

#ws['D2'] = gene
#ws['B2'] = condition[gene]
#ws['E10'] = website[gene]
#ws['D3'] = var
#ws['F8'] = locus

# Save the file
#output_path = sys.argv[1] + "_variant_worksheet.xlsx" 
#wb.save(output_path)

bb_gene = gene
bb_var = var
bb_query = bb_gene + ":" + bb_var
bb_condition = "Results unavailable"
bb_hgvsp = "Results unavailable"
bb_chrom = "Results unavailable" 
bb_start_end = "Results unavailable"
bb_locus = "Results unavailable"
bb_gnomad = "Results unavailable"
bb_exac = "Results unavailable"
bb_dbsnp = "Results unavailable" 
bb_clinvar = "Results unavailable"
bb_splice = "Results unavailable"
bb_ucsc = "Results unavailable"
bb_polyphen = "Results unavailable"
bb_cadd = "Results unavailable"
bb_sift = "Results unavailable"
bb_provean = "Results unavailable"
bb_user = sys.argv[5]
bb_date = sys.argv[6]

print("STARTING DATA IMPORT")

command = "python3.7 /home/klong/bb_django/bluebonnet/import_data.py" + " '" + bb_query + "' '" + bb_condition + "' '" + bb_hgvsp + "' '" + bb_chrom + "' '" + bb_start_end + "' '" + bb_locus + "' '" + bb_gnomad + "' '" + bb_exac + "' '" + bb_dbsnp + "' '" + bb_clinvar + "' '" + bb_splice + "' '" + bb_ucsc + "' '" + bb_polyphen + "' '" + bb_cadd + "' '" + bb_sift + "' '" + bb_provean + "' '" + bb_date + "' '" + bb_user + "'"
os.system(command)
