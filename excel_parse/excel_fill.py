from openpyxl import load_workbook
import os
import sys
import re
import statistics
import subprocess

base_dir = "/work/NBS/"

#Grab template excel template, set worksheet
#workbook_path = base_dir + "/excel_parse/variant_worksheet_v2.xlsx"
#wb = load_workbook(workbook_path)
#ws = wb.active

#Output from pipeline
input_file=sys.argv[1] + ".all"
clinvar_file=sys.argv[1] + "_clinvar.txt"
provean_file=sys.argv[1] + "_provean.txt"
fsplice_file=sys.argv[1] + "_fsplice.txt"
dbSNP_freqs_file=sys.argv[1] + "_dbSNP_freqs.txt"
gnomad_file=sys.argv[1] + "_gnomad.txt"
full_out = open(input_file,"r").readlines()

#initialize some vars
at_ucsc = False
ucsc_out = []
atPPH2 = False
pph2_full = ""
atMAF = False
MAFout = ""
atSplice = False
spliceOut = ""
#Values based on queue, not outputs

gene=sys.argv[3]
condition = {
  "ABCD1": "X-ALD",
  "HBB": "Hemoglobinopathy",
  "ACADVL": "VLCADD"
}
website = {
  "ABCD1": "https://adrenoleukodystrophy.info/mutations-and-variants-in-abcd1",
  "HBB": "https://globin.bx.psu.edu/cgi-bin/hbvar/query_vars3",
  "ACADVL": "N/A"
}



#$prefix $rsid $gene $var $locus $pos $end $chrom
rsid=sys.argv[2]
bb_rsid = rsid
var=sys.argv[4]
bb_var = var
locus=sys.argv[5]
bb_locus = locus
start=sys.argv[6]
bb_start = start
end=sys.argv[7]
bb_end = end
chrom=sys.argv[8]
bb_chrom = chrom
ref=sys.argv[9]
bb_ref = ref
alt=sys.argv[10]
bb_alt = alt
transcript=sys.argv[11]
bb_transcript = transcript
hgvsp=sys.argv[12]
bb_hgvsp = hgvsp
bb_user=sys.argv[13]
bb_date=sys.argv[14]


#ws['H3'] = gene
#ws['D2'] = gene
bb_gene = gene
#ws['B2'] = condition[gene]
bb_condition = condition[gene]
#ws['D3'] = var
bb_var = var
#ws['F8'] = locus
bb_locus = locus
#ws['B8'] = chrom
#ws['D8'] = start + " - " + end
bb_start_end = start + " - " + end
#ws['H3'] = hgvsp

for i,line in enumerate(full_out):
    
    #gnomad
    if re.search("gnomad.txt",line):
        if re.search("Mutation not found",full_out[i+1]):
            #ws['D20'] = "Mutation not found within gnomad"
            bb_gnomad = "Mutation not found within gnomad"
        else:
            af = str(round(float(full_out[i+8][:-2].split()[1])*100,4)) + "%\n"
            af = "Overall Allele Frequency: " + af 
            gnomad_freq = af
            for x in range(9, 20):
                if full_out[i+x].startswith('{"id":'):
                    group = full_out[i+x].split('"')[3]
                    ac = full_out[i+x].split('"')[6].split(" ")[1].replace(",","")
                    an = full_out[i+x].split('"')[8].split(" ")[1].replace("},","").replace("}]}}","").strip()
                    perc = round(float(float(ac)/float(an))*100,4)
                    pop_append = group + ": " + str(perc) + "% (" + ac + "/" + an + ")\n"
                    gnomad_freq = gnomad_freq + pop_append
            #ws['D20'] = gnomad_freq
            bb_gnomad = gnomad_freq
    #exac
    if re.search("exac.txt",line):
        if re.search("Mutation not found",full_out[i+1]):
            #ws['D15'] = "Mutation not found within exac"
            bb_exac = "Mutation not found within exac"
        else:
            af = str(round(float(full_out[i+8][:-2].split()[1])*100,4)) + "%\n"
            af = "Overall Allele Frequency: " + af
            exac_freq = af
            for x in range(9, 20):
                if full_out[i+x].startswith('{"id":'):
                    group = full_out[i+x].split('"')[3]
                    ac = full_out[i+x].split('"')[6].split(" ")[1].replace(",","")
                    an = full_out[i+x].split('"')[8].split(" ")[1].replace("},","").replace("}]}}","").strip()
                    perc = round(float(float(ac)/float(an))*100,4)
                    pop_append = group + ": " + str(perc) + "% (" + ac + "/" + an + ")\n"
                    exac_freq = exac_freq + pop_append
            #ws['D15'] = exac_freq
            bb_exac = exac_freq

    #dbSNP rsID
    if rsid == "NA":
        #ws['D25'] = "This query does not have a dbSNP entry"
        bb_rsid = "This query does not have a dbSNP entry"
        MAFout = "N/A"
    else:
        #ws['D25'] = rsid
        bb_rsid = rsid

    #dbSNP frequencies
    '''
    if atMAF == True:
        if line.startswith("*"):
            atMAF = False
        elif line.strip()=="":
            continue
        elif " NA " in line:
            MAFout = "dbSNP allele frequencies were unavailable for this mutation"
            atMAF = False
        else:
            MAFout = MAFout + line.split()[1] + "   " + str(round(float(line.split()[-1])*100,2)) + "%\n"
    if "study" in line and "MAF" in line:
        atMAF = True
        MAFout = "study   MAF\n"
    '''

    #CADD
    if re.search("CADD.txt",line):
        if full_out[i+3].strip() == "":
            #ws['E48'] = "CADD results unavailable"
            bb_cadd = "CADD results unavailable"
        else:
            CADD_result = full_out[i+3].split("\t")
            #ws['E48'] = "PHRED: " + CADD_result[5].strip()
            bb_cadd = "PHRED: " + CADD_result[5].strip()

    '''
    #intervar
    query_by = "Chr: " + chrom + ", Start: " + start + ", End: " + end + ", Ref: " + ref + ", Alt: " + alt 
    intervar_status = ["xpathogenic", "likely pathogenic", "uncertain significance", "likely benign", "xbenign"]
    if re.search("intervar.txt",line):
        intervar_results = full_out[i+2].split("\t")[13].replace("InterVar: ","x")
        for status in intervar_status:
            if re.search(status, intervar_results, re.IGNORECASE):
                ws['E40'] = query_by
                ws['E41'] = status.replace("x","")
                intervar_details = re.split(status, intervar_results, flags=re.IGNORECASE)[1].strip().replace(", ",",").split()
                details_out = ""
                for i in intervar_details:
                    detail_prefix = i.split("=")[0]
                    #print(detail_prefix)
                    if "[" in i.split("=")[1]:
                        detail_result = i.split("=")[1].strip('][').split(',')
                        #print(detail_result)
                        count = 0
                        for x in detail_result:
                            #print(x)
                            count = count + 1
                            if x == "1":
                                details_out = details_out + detail_prefix + str(count) + "  "
                    else:
                        if i.split("=")[1] == "1":
                            details_out = details_out + detail_prefix + " "
                ws['E42'] = details_out
    '''     

    #UCSC, phylop100way
    if re.search("UCSC.txt",line):
        at_ucsc=True
    if at_ucsc == True and line.startswith("{u"):
        ucsc_out.append(float(line.split(": ")[3].replace("}","").strip()))
    if at_ucsc == True and line.startswith("*"):
        at_ucsc = False
        ucsc_mean = round(statistics.mean(ucsc_out),2)
        ucsc_min = round(min(ucsc_out),2)
        ucsc_max = round(max(ucsc_out),2)
        #ws['E46'] = "Average Conservation Score: " + str(ucsc_mean) + "\nMin: " + str(ucsc_min) + "\nMax: " + str(ucsc_max)
        bb_ucsc = "Average Conservation Score: " + str(ucsc_mean) + "\nMin: " + str(ucsc_min) + "\nMax: " + str(ucsc_max)

    #SIFT indel
    if re.search("SIFTindel.txt",line):
        if full_out[i+2].strip() == "":
            #ws['E49'] = "SIFT results unavailable"
            bb_sift = "SIFT results unavailable"
        else:
            SIFT_result = full_out[i+2].split("\t")
            if SIFT_result[6]=="" and SIFT_result[7]=="":
                #ws['E49'] = "SIFT results unavailable"
                bb_sift = "SIFT results unavailable"
            else:
                #ws['E49'] = "Effect: " + SIFT_result[6] + "\nConfidence Score: " + SIFT_result[7] 
                bb_sift = "Effect: " + SIFT_result[6] + "\nConfidence Score: " + SIFT_result[7]

    #SIFT snv
    if re.search("SIFT.txt",line):
        sift_out = ""
        if full_out[i+8] == "":
            #ws['E49'] = "SIFT results unavailable"
            bb_sift = "SIFT results unavailable"
        else:
            SIFT_result = full_out[i+8].split("\t")[7]
            if re.search("SIFTINFO=",SIFT_result):
                sift_info = SIFT_result.split("SIFTINFO=")[1]
                try:
                    sift_info_list = sift_info.split(',')
                    for entry in sift_info_list:
                        if entry.split("|")[1] == transcript:
                            sift_out = "SIFT Score: " + entry.split("|")[8] + "\nSIFT Prediction: " + entry.split("|")[12]
                    if sift_out == "":
                        sift_out = "Transcript not found in SIFT output"
                except ValueError:
                    if sift_info.split("|")[1] == transcript:
                        sift_out = "SIFT Score: " + entry.split("|")[8] + "\nSIFT Prediction: " + entry.split("|")[12]
                    else:
                        sift_out = ""
                #ws['E49'] = sift_out
                bb_sift = sift_out
            else:
                #ws['E49'] = "SIFT results unavailable"
                bb_sift = "SIFT results unavailable"

    #polyphen
    if atPPH2 == True:
        if line.startswith("*"):
            atPPH2 = False
        elif "PolyPhen=" in line and line.split()[4] == transcript:
            pph2_full = pph2_full + line.split("PolyPhen=")[1].strip() + ", "
        else:
            continue
    if re.search("polyphen2.txt",line):
        atPPH2 = True

#clinvar
clinvar_excel_out = ""
clinvar_out = open(clinvar_file,"r").readlines()
for i in clinvar_out:
    clinvar_excel_out = clinvar_excel_out + i + "\n"
#ws['B32'] = str(clinvar_excel_out)
bb_clinvar = str(clinvar_excel_out)

#polyphen
if pph2_full == "":
    #ws['E47'] = "Polyphen results unavailable for this query"
    bb_polyphen = "Polyphen results unavailable for this query"
else:
    pph2_score = pph2_full.split("(")[1][:-3]
    pph2_pred = pph2_full.split("(")[0]
    #ws['E47'] = "Score: " + pph2_score + "\nPrediction: " + pph2_pred
    bb_polyphen = "Score: " + pph2_score + "\nPrediction: " + pph2_pred
#ws['D26'] = MAFout.strip()

#provean
provean_excel_out = ""
provean_out = open(provean_file,"r").readlines()
provean_excel_out = provean_out[-1].strip()
if provean_excel_out.startswith("N/A:"):
    #ws['E50'] = provean_excel_out
    bb_provean = provean_excel_out
else:
    provean_query = provean_excel_out.split("\t")[0]
    provean_score = provean_excel_out.split("\t")[1]
    #ws['E50'] = "Query: " + provean_query + "\nScore: " + provean_score
    bb_provean = "Query: " + provean_query + "\nScore: " + provean_score  

#splice
fsplice_out = open(fsplice_file,"r").read()
#ws['D36'] = fsplice_out.strip()
bb_splice = fsplice_out.strip()

#dbsnp frequencies
dbSNP_freqs_excel_out = ""
dbSNP_freqs_out = open(dbSNP_freqs_file,"r").readlines()
#print(dbSNP_freqs_out)
for i in dbSNP_freqs_out:
    #print(i)
    freq_split = i.split("\t")
    #print(freq_split)
    study = freq_split[0] + "." + freq_split[1]
    mut = freq_split[2] + "/" + freq_split[3]
    count = freq_split[4]
    total = freq_split[5].strip()
    perc = round(float(float(count)/float(total))*100,4)
    add_line = study + ": " + str(perc) + "% (" + count + "/" + total + ")\n"
    dbSNP_freqs_excel_out = dbSNP_freqs_excel_out + add_line
#ws['B27'] = dbSNP_freqs_excel_out
bb_dbsnp = dbSNP_freqs_excel_out
bb_dbsnp = "RSID: " + bb_rsid + "\n" + bb_dbsnp

# Save the file
#output_path = sys.argv[1] + "_variant_worksheet.xlsx" 
#wb.save(output_path)

#add to entry to db
bb_query = bb_gene + ":" + bb_var
command = "python3.7 /home/klong/bb_django/bluebonnet/import_data.py" + " '" + bb_query + "' '" + bb_condition + "' '" + bb_hgvsp + "' '" + bb_chrom + "' '" + bb_start_end + "' '" + bb_locus + "' '" + bb_gnomad + "' '" + bb_exac + "' '" + bb_dbsnp + "' '" + bb_clinvar + "' '" + bb_splice + "' '" + bb_ucsc + "' '" + bb_polyphen + "' '" + bb_cadd + "' '" + bb_sift + "' '" + bb_provean + "' '" + bb_date + "' '" + bb_user + "'"
os.system(command)
